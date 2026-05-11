from django.contrib import admin , messages
from django.core.management import call_command
from mptt.admin import MPTTModelAdmin ,DraggableMPTTAdmin
from .models import (OrganizationalStructure , Job , JobTitle ,Employee , Shift , Attendance,Day,EducationalLevel , TypeOfEmployee , MaritalStatus , DeviceFigerPrint , EmployeeFingerPrint , DataResptions ,LawFingerPrinter)
from .models import AttendanceStatus
from django.utils import timezone
from datetime import date, timedelta ,datetime
from django.shortcuts import render
from django.http import HttpResponse
import openpyxl
from django.db.models import Q
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
import calendar
from .forms import MonthlyReportForm
from rangefilter.filters import DateRangeFilter
# # Register your models here.

def create_excel_report(queryset, year, month):
    status_display_map = dict(AttendanceStatus.choices)

    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = f"تقرير شهر {month}-{year}"
    ws.sheet_view.rightToLeft = True
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    weekend_fill = PatternFill(start_color="DCDCDC", fill_type="solid")

    sub_headers = ["دخول", "خروج", "الحالة"]
    ws.cell(row=1, column=1, value="اسم الموظف").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    _, num_days = calendar.monthrange(year, month)
    col_idx = 2
    for day in range(1, num_days + 1):
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 2)
        cell = ws.cell(row=1, column=col_idx, value=day)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align
        for i, sub_header in enumerate(sub_headers):
            cell = ws.cell(row=2, column=col_idx + i, value=sub_header)
            cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align
        col_idx += 3

    summary_headers = ["أيام الدوام", "أيام الغياب", "إجمالي التأخير"]
    summary_col_start = col_idx
    for i, header in enumerate(summary_headers):
        ws.merge_cells(start_row=1, start_column=summary_col_start + i, end_row=2, end_column=summary_col_start + i)
        cell = ws.cell(row=1, column=summary_col_start + i, value=header)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align

    start_date = date(year, month, 1)
    end_date = date(year, month, num_days)
    row_idx = 3

    for emp in queryset:
        ws.cell(row=row_idx, column=1, value=emp.name)

        emp_shifts = Shift.objects.filter(employee=emp)
        # --- [تعديل 1: إنشاء خريطة للدوامات لسهولة الوصول] ---
        # هذا يجعل الكود أسرع بدلاً من البحث في كل مرة داخل الحلقة
        shift_map = {day.day: shift for shift in emp_shifts for day in shift.days.all()}

        attendance_dict = {rec.date: rec for rec in Attendance.objects.filter(employee=emp, date__range=(start_date, end_date))}

        total_present, total_absent, total_lateness = 0, 0, timedelta()
        col_idx = 2

        for day in range(1, num_days + 1):
            current_date = date(year, month, day)
            day_name_map = {0: 'mon', 1: 'tue', 2: 'wed', 3: 'thu', 4: 'fri', 5: 'sat', 6: 'sun'}
            day_str = day_name_map[current_date.weekday()]

            # الحصول على الشفت الخاص بهذا اليوم
            relevant_shift = shift_map.get(day_str)

            if not relevant_shift: # إذا لم يكن هناك شفت لهذا اليوم، فهو يوم عطلة
                ws.merge_cells(start_row=row_idx, start_column=col_idx, end_row=row_idx, end_column=col_idx + 2)
                ws.cell(row=row_idx, column=col_idx, value="عطلة").fill = weekend_fill
            else:
                att_record = attendance_dict.get(current_date)
                if att_record:
                    ws.cell(row=row_idx, column=col_idx, value=att_record.time_in.strftime('%H:%M') if att_record.time_in else '-')
                    ws.cell(row=row_idx, column=col_idx + 1, value=att_record.time_out.strftime('%H:%M') if att_record.time_out else '-')
                    status_text = status_display_map.get(att_record.is_present1, "غير معروف")
                    ws.cell(row=row_idx, column=col_idx + 2, value=status_text)

                    # --- [تعديل 2: الجزء الأهم - حساب التأخير] ---
                    if att_record.time_in and relevant_shift.start_time:
                        # لكي نتمكن من طرح الوقت، يجب تحويلهما إلى كائن datetime كامل
                        scheduled_start_dt = datetime.combine(current_date, relevant_shift.start_time)
                        actual_arrival_dt = datetime.combine(current_date, att_record.time_in)

                        # إذا وصل الموظف بعد وقت الدوام الرسمي
                        if actual_arrival_dt > scheduled_start_dt:
                            lateness_this_day = actual_arrival_dt - scheduled_start_dt
                            total_lateness += lateness_this_day # هنا نضيف التأخير للمجموع
                    # --- [نهاية تعديل 2] ---

                    if att_record.is_present1 not in [AttendanceStatus.ABSENT, AttendanceStatus.VACATION]: total_present += 1
                    else: total_absent += 1
                else: # حالة عدم وجود سجل حضور في يوم عمل
                    ws.cell(row=row_idx, column=col_idx, value='-')
                    ws.cell(row=row_idx, column=col_idx + 1, value='-')
                    status_text = status_display_map.get(AttendanceStatus.ABSENT, "غائب")
                    ws.cell(row=row_idx, column=col_idx + 2, value=status_text)
                    total_absent += 1
            col_idx += 3

        ws.cell(row=row_idx, column=summary_col_start, value=total_present)
        ws.cell(row=row_idx, column=summary_col_start + 1, value=total_absent)

        # هذا الجزء سيعمل الآن بشكل صحيح لأن total_lateness تحتوي على قيمة
        total_seconds = total_lateness.total_seconds()
        hours, remainder = divmod(total_seconds, 3600)
        minutes, _ = divmod(remainder, 60)
        ws.cell(row=row_idx, column=summary_col_start + 2, value=f"{int(hours):02}:{int(minutes):02}")
        row_idx += 1

    ws.column_dimensions[get_column_letter(1)].width = 25
    for i in range(2, ws.max_column + 1): ws.column_dimensions[get_column_letter(i)].width = 15

    return workbook







# def create_excel_report(queryset, year, month):
#     status_display_map = dict(AttendanceStatus.choices)

#     workbook = openpyxl.Workbook()
#     ws = workbook.active
#     ws.title = f"تقرير شهر {month}-{year}"
#     ws.sheet_view.rightToLeft = True
#     header_font = Font(bold=True, color="FFFFFF")
#     header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
#     center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
#     weekend_fill = PatternFill(start_color="DCDCDC", fill_type="solid")

#     sub_headers = ["دخول", "خروج", "الحالة"]
#     ws.cell(row=1, column=1, value="اسم الموظف").font = header_font
#     ws.cell(row=1, column=1).fill = header_fill
#     ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

#     _, num_days = calendar.monthrange(year, month)
#     col_idx = 2
#     for day in range(1, num_days + 1):
#         ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 2)
#         cell = ws.cell(row=1, column=col_idx, value=day)
#         cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align
#         for i, sub_header in enumerate(sub_headers):
#             cell = ws.cell(row=2, column=col_idx + i, value=sub_header)
#             cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align
#         col_idx += 3

#     summary_headers = ["أيام الدوام", "أيام الغياب", "إجمالي التأخير"]
#     summary_col_start = col_idx
#     for i, header in enumerate(summary_headers):
#         ws.merge_cells(start_row=1, start_column=summary_col_start + i, end_row=2, end_column=summary_col_start + i)
#         cell = ws.cell(row=1, column=summary_col_start + i, value=header)
#         cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align

#     start_date = date(year, month, 1)
#     end_date = date(year, month, num_days)
#     row_idx = 3

#     for emp in queryset:
#         ws.cell(row=row_idx, column=1, value=emp.name)

#         emp_shifts = Shift.objects.filter(employee=emp)
#         work_days_str = {day.day for shift in emp_shifts for day in shift.days.all()}
#         attendance_dict = {rec.date: rec for rec in Attendance.objects.filter(employee=emp, date__range=(start_date, end_date))}

#         total_present, total_absent, total_lateness = 0, 0, timedelta()
#         col_idx = 2

#         for day in range(1, num_days + 1):
#             current_date = date(year, month, day)
#             day_str = {0: 'mon', 1: 'tue', 2: 'wed', 3: 'thu', 4: 'fri', 5: 'sat', 6: 'sun'}[current_date.weekday()]
#             is_workday = day_str in work_days_str

#             if not is_workday:
#                 ws.merge_cells(start_row=row_idx, start_column=col_idx, end_row=row_idx, end_column=col_idx + 2)
#                 ws.cell(row=row_idx, column=col_idx, value="عطلة").fill = weekend_fill
#             else:
#                 att_record = attendance_dict.get(current_date)
#                 if att_record:
#                     ws.cell(row=row_idx, column=col_idx, value=att_record.time_in.strftime('%H:%M') if att_record.time_in else '-')
#                     ws.cell(row=row_idx, column=col_idx + 1, value=att_record.time_out.strftime('%H:%M') if att_record.time_out else '-')
#                     status_text = status_display_map.get(att_record.is_present1, "غير معروف")
#                     ws.cell(row=row_idx, column=col_idx + 2, value=status_text)

#                     if att_record.is_present1 not in [AttendanceStatus.ABSENT, AttendanceStatus.VACATION]: total_present += 1
#                     else: total_absent += 1
#                 else:
#                     ws.cell(row=row_idx, column=col_idx, value='-')
#                     ws.cell(row=row_idx, column=col_idx + 1, value='-')
#                     status_text = status_display_map.get(AttendanceStatus.ABSENT, "غائب")
#                     ws.cell(row=row_idx, column=col_idx + 2, value=status_text)

#                     total_absent += 1
#             col_idx += 3
#         ws.cell(row=row_idx, column=summary_col_start, value=total_present)
#         ws.cell(row=row_idx, column=summary_col_start + 1, value=total_absent)
#         total_seconds = total_lateness.total_seconds()
#         hours, remainder = divmod(total_seconds, 3600)
#         minutes, _ = divmod(remainder, 60)
#         ws.cell(row=row_idx, column=summary_col_start + 2, value=f"{int(hours):02}:{int(minutes):02}")
#         row_idx += 1

#     ws.column_dimensions[get_column_letter(1)].width = 25
#     for i in range(2, ws.max_column + 1): ws.column_dimensions[get_column_letter(i)].width = 15

#     return workbook

@admin.action(description="تصدير تقرير الحضور الشهري")
def export_monthly_report_action(modeladmin, request, queryset):
    if 'post' in request.POST:
        form = MonthlyReportForm(request.POST)
        if form.is_valid():
            month = int(form.cleaned_data['month'])
            year = int(form.cleaned_data['year'])
            workbook = create_excel_report(queryset, year, month)
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="attendance_report_{year}_{month}.xlsx"'
            workbook.save(response)
            return response
    else:
        form = MonthlyReportForm()

    return render(request, 'admin/hr/employee/monthly_report_intermediate.html', {
        'queryset': queryset,
        'form': form,
        'action_checkbox_name': admin.helpers.ACTION_CHECKBOX_NAME,
    })

@admin.action(description='تحديث الحضور للموظفين المحددين (لليوم)')
def update_selected_employees_attendance(self, request, queryset):
    employee_ids = list(queryset.values_list('id', flat=True))

    if not employee_ids:
        self.message_user(request, "لم يتم تحديد أي موظف.", messages.WARNING)
        return

    today_str = timezone.now().date().strftime('%Y-%m-%d')

    try:
        call_command('process_attendance', date=today_str, employees=employee_ids)
        self.message_user(request, f"تم تحديث بيانات الحضور لـ {len(employee_ids)} موظف بنجاح.", messages.SUCCESS)
    except Exception as e:
        self.message_user(request, f"حدث خطأ: {e}", messages.ERROR)

@admin.action(description='تحديث الحضور لجميع الموظفين (لليوم)')
def update_all_employees_for_today(self, request, queryset):
    today_str = timezone.now().date().strftime('%Y-%m-%d')

    try:
        call_command('process_attendance', date=today_str)
        self.message_user(request, "تم بدء عملية تحديث بيانات الحضور لجميع الموظفين بنجاح.", messages.SUCCESS)
    except Exception as e:
        self.message_user(request, f"حدث خطأ: {e}", messages.ERROR)



@admin.register(TypeOfEmployee)
class TypeOfEmployeeAdmin(admin.ModelAdmin):
    list_display = ('type_of_employee',)

class EmployeeFilter(admin.SimpleListFilter):
    title = 'الموظف'  # اسم الفلتر الذي سيظهر في القائمة
    parameter_name = 'employee'  # اسم المتغير في رابط URL

    def lookups(self, request, model_admin):
        # هذه الدالة تنشئ قائمة الخيارات للفلتر
        # نحن نجلب فقط الموظفين الذين لديهم بصمات مسجلة
        employees = Employee.objects.filter(employeefingerprint__isnull=False).distinct()
        return [(e.id, e.name) for e in employees]

    def queryset(self, request, queryset):
        # هذه الدالة تطبق الفلترة على البيانات
        if self.value():
            # إذا اختار المستخدم موظفاً، نحصل على كل أرقام البصمة الخاصة به
            user_ids = EmployeeFingerPrint.objects.filter(employee_id=self.value()).values_list('id_users', flat=True)
            # ثم نفلتر سجلات البصمات بناءً على هذه الأرقام
            return queryset.filter(user_id__in=user_ids)
        return queryset

@admin.register(DataResptions)
class DataResptionsAdmin(admin.ModelAdmin):
    # 1. الحقول التي ستظهر في القائمة
    list_display = ('employee_name', 'device_finger_print', 'timestamp', 'status','create_at')

    # 2. الفلاتر التي ستظهر على الجانب
    list_filter = ('timestamp', 'device_finger_print', ('timestamp', DateRangeFilter)) # أضفنا فلتر التاريخ وفلتر الموظف

    # 3. حقول البحث
    search_fields = ('user_id',) # البحث الأساسي برقم البصمة

    # دالة مخصصة لجلب وعرض اسم الموظف
    @admin.display(description='اسم الموظف', ordering='-timestamp')
    def employee_name(self, obj):
        # نبحث عن الموظف المرتبط برقم البصمة (user_id)
        fingerprint = EmployeeFingerPrint.objects.filter(id_users=obj.user_id).first()
        if fingerprint:
            return fingerprint.employee.name
        return f"غير معروف ({obj.user_id})"

    # دالة مخصصة لتفعيل البحث باسم الموظف (ميزة متقدمة)
    def get_search_results(self, request, queryset, search_term):
        queryset, use_distinct = super().get_search_results(request, queryset, search_term)
        if search_term:
            # نبحث عن أرقام البصمة للموظفين الذين يتطابق اسمهم مع البحث
            matching_user_ids = EmployeeFingerPrint.objects.filter(
                employee__name__icontains=search_term
            ).values_list('id_users', flat=True)

            # ندمج نتائج البحث بالاسم مع نتائج البحث بالرقم باستخدام OR
            queryset |= self.model.objects.filter(user_id__in=matching_user_ids)

        return queryset, use_distinct



@admin.register(MaritalStatus)
class MaritalStatusAdmin(admin.ModelAdmin):
    list_display = ('marital_status',)

@admin.register(EducationalLevel)
class EducationalLevelAdmin(admin.ModelAdmin):
    list_display = ('name',)

@admin.register(Day)
class DayAdmin(admin.ModelAdmin):
    list_display =('day',)

@admin.register(OrganizationalStructure)
class OrganizationalStructureAdmin(DraggableMPTTAdmin):
    list_display = ('tree_actions','indented_title','name',)
    # list_display_links ('indented_title',)

# admin.site.register(
#     OrganizationalStructure,
#     DraggableMPTTAdmin,
#     list_display=(
#         'tree_actions',
#         'indented_title',
#         'name'
#     ),
#     list_display_links=(
#         'indented_title',
#     ),
# )

class JobInline(admin.TabularInline):
    model = Job
    extra = 1

class EmployeeFingerPrintInline(admin.TabularInline):
    model = EmployeeFingerPrint
    extra = 1
@admin.register(Job)
class JobAdmin(admin.ModelAdmin):
    list_display = ('employee','organizational_structure','JobTitle' ,'create_at')
    list_filter = ('employee','organizational_structure','JobTitle')

@admin.register(Employee)
class EmployeeAdmin(admin.ModelAdmin):
    list_display = ('name','educational_level','type_of_employee','phone','address')
    list_filter = ('active','type_of_employee','educational_level')
    search_fields =('name',)
    inlines = [JobInline , EmployeeFingerPrintInline]
    actions = [export_monthly_report_action , update_all_employees_for_today]

@admin.register(DeviceFigerPrint)
class DeviceFigerPrintAdmin(admin.ModelAdmin):
    list_display = ('name','location')

@admin.register(JobTitle)
class JobTitleAdmin(admin.ModelAdmin):
    list_display = ('name','tesk')

@admin.register(Shift)
class ShiftAdmin(admin.ModelAdmin):
    list_display = ('name', 'start_time', 'end_time',)
    search_fields = ('name',)



@admin.register(Attendance)
class AttendanceAdmin(admin.ModelAdmin):
    list_display = ('employee', 'time_in', 'time_out', 'is_present', 'date', 'note')
    list_filter = (('create_at', DateRangeFilter), 'employee', 'is_present', 'date')


@admin.register(LawFingerPrinter)
class LawFingerPrinterAdmin(admin.ModelAdmin):
    list_display = ('name','shift','entry_grace_period','consider_absent_if_late_by','early_departure_allowance')
    list_filter = ('shift',)




